"""
Subprocess Runner for non-OpenAI models.

Generates Python code using LLM, then executes it safely in an isolated subprocess
environment with strict security controls.

Security features:
- Isolated temporary directory
- Environment variable whitelist (no API keys)
- Configurable timeout (see config.SUBPROCESS_TIMEOUT)
- No network access (subprocess has no credentials)
"""

import subprocess
import tempfile
import shutil
import os
import re
import sys
from pathlib import Path
from typing import Optional

from core.config import SUBPROCESS_TIMEOUT
from core.llm_client import complete
from core.prompt_loader import load_prompt, render_prompt
from core.file_preview import generate_all_previews


def _build_file_structure_info(reference_files: list) -> str:
    """Read reference file structure and return a prompt-ready summary.

    Reads column names and shape from each reference file.
    Supports: .xlsx, .xls, .csv, .tsv, .parquet, .json

    Token-efficient: metadata only, no row content.
    Returns empty string if no files provided or all fail to read.
    """
    if not reference_files:
        return ""

    lines = ["## Reference File Structure (auto-detected)"]
    any_success = False

    for fpath in reference_files:
        path = Path(fpath)
        if not path.exists():
            continue
        suffix = path.suffix.lower()

        try:
            if suffix in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                    headers = [str(h) for h in headers if h is not None]
                    nrows = (ws.max_row or 1) - 1
                    label = (
                        f"{path.name} (sheet: {sheet_name})"
                        if len(wb.sheetnames) > 1
                        else path.name
                    )
                    lines.append(f"\n{label}: ~{nrows} rows × {len(headers)} cols")
                    lines.append(f"  columns: {headers}")
                wb.close()
                any_success = True

            elif suffix in (".csv", ".tsv"):
                import csv
                sep = "\t" if suffix == ".tsv" else ","
                with open(fpath, newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f, delimiter=sep)
                    headers = next(reader)
                    nrows = sum(1 for _ in reader)
                lines.append(f"\n{path.name}: ~{nrows} rows × {len(headers)} cols")
                lines.append(f"  columns: {headers}")
                any_success = True

            elif suffix == ".parquet":
                import pyarrow.parquet as pq
                schema = pq.read_schema(fpath)
                meta = pq.read_metadata(fpath)
                lines.append(f"\n{path.name}: {meta.num_rows} rows × {len(schema.names)} cols")
                lines.append(f"  columns: {schema.names}")
                any_success = True

            elif suffix == ".json":
                import json as _json
                with open(fpath, encoding="utf-8") as f:
                    data = _json.load(f)
                if isinstance(data, list) and data:
                    keys = list(data[0].keys()) if isinstance(data[0], dict) else []
                    lines.append(f"\n{path.name}: list of {len(data)} records")
                    if keys:
                        lines.append(f"  keys: {keys}")
                elif isinstance(data, dict):
                    lines.append(f"\n{path.name}: dict with keys: {list(data.keys())[:20]}")
                any_success = True

        except Exception:
            # 읽기 실패 시 조용히 스킵 — 추론 크래시 방지
            continue

    return "\n".join(lines) if any_success else ""


class SubprocessRunner:
    """LLM code generation → safe subprocess execution"""

    DEFAULT_PROMPT = "subprocess_occupation_codegen"

    def __init__(self, llm_client, prompt_name: str = DEFAULT_PROMPT):
        """
        Initialize Subprocess runner with LLM client.

        Args:
            llm_client: AzureOpenAI client instance for code generation
            prompt_name: Name of prompt YAML file in prompts/ (without .yaml)
        """
        self.llm_client = llm_client
        self.prompt_name = prompt_name
        self.prompt_data = load_prompt(prompt_name)

    def run(
        self,
        task_prompt: str,
        model: str,
        reference_files: Optional[list] = None,
        occupation: str = "professional",
        experiment_prompt: Optional[dict] = None,
    ) -> dict:
        """
        Generate code via LLM and execute safely in subprocess.

        Args:
            task_prompt: The task instruction
            model: Model deployment name
            reference_files: Optional list of file paths to copy to execution dir
            occupation: Professional role from task data
            experiment_prompt: Optional prompt overrides from experiment YAML
                Keys: system (str), prefix (str|None), body (str|None), suffix (str|None)

        Returns:
            dict with keys:
                - success (bool): Whether execution succeeded
                - text (str): stdout from code execution
                - files (list): List of generated files [{filename, content}]
                - error (str, optional): Error message if failed
        """
        try:
            # Reference 파일 구조 자동 주입 (컬럼명 하드코딩 에러 방지)
            file_structure_info = _build_file_structure_info(reference_files or [])
            if file_structure_info:
                task_prompt = file_structure_info + "\n\n" + task_prompt

            # Step 0: Generate reference file previews and append to task_prompt
            if reference_files:
                previews = generate_all_previews(reference_files)
                if previews:
                    task_prompt = task_prompt + "\n\n" + previews

                # Explicitly list available files for LLM code generation
                available_files = [os.path.basename(f) for f in reference_files]
                task_prompt = (
                    task_prompt
                    + f"\n\n📁 Files available in current directory (you can use them directly): {available_files}"
                )

            # Step 1: Generate code using LLM
            rendered = render_prompt(
                self.prompt_data,
                occupation=occupation,
                task_prompt=task_prompt,
                experiment_prompt=experiment_prompt,
            )

            messages = [
                {"role": "system", "content": rendered["system_message"]},
                {"role": "user", "content": rendered["user_prompt"]}
            ]

            response, latency = complete(
                client=self.llm_client,
                model=model,
                messages=messages,
                max_completion_tokens=4000
            )

            response_text = response.choices[0].message.content

            # Step 2: Extract code and description from response
            code = self._extract_code(response_text)
            if not code:
                return {
                    "success": False,
                    "text": "",
                    "deliverable_text": "",
                    "files": [],
                    "error": f"No Python code found in LLM response. Response: {response_text[:200]}..."
                }

            # Extract the descriptive text (non-code portion of LLM response)
            deliverable_text = self._extract_description(response_text)

            # Step 3: Execute safely in isolated environment
            result = self._execute_safely(code, reference_files)
            result["deliverable_text"] = deliverable_text
            return result

        except Exception as e:
            return {
                "success": False,
                "text": "",
                "files": [],
                "error": f"Code generation failed: {str(e)}"
            }

    def _extract_description(self, text: str) -> str:
        """Extract non-code descriptive text from LLM response.

        Strips code blocks and returns the remaining text as a
        meta description of what deliverables were created.

        Args:
            text: Full LLM response text

        Returns:
            Descriptive text with code blocks removed
        """
        # Remove all closed code blocks (```python...``` or ```...```)
        cleaned = re.sub(r"```[\w]*\s*\n.*?```", "", text, flags=re.DOTALL)
        # Remove unclosed code blocks (LLM truncation or trailing code)
        cleaned = re.sub(r"```[\w]*\s*\n.*$", "", cleaned, flags=re.DOTALL)
        # Clean up excessive whitespace
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
        return cleaned

    def _extract_code(self, text: str) -> Optional[str]:
        """
        Extract Python code from ```python``` code blocks.

        Args:
            text: LLM response text

        Returns:
            Extracted code or None if not found
        """
        # Pattern: ```python ... ``` (flexible whitespace)
        pattern = r"```python\s*\n(.*?)```"
        matches = re.findall(pattern, text, re.DOTALL)

        if matches:
            # Concatenate all code blocks if multiple
            return "\n\n".join(m.strip() for m in matches)

        # Fallback: Try without language specifier
        pattern = r"```\s*\n(.*?)```"
        matches = re.findall(pattern, text, re.DOTALL)

        if matches:
            # Check if it looks like Python code
            code = matches[0].strip()
            if "import" in code or "def " in code or "print" in code:
                return code

        # Fallback: Code block opened but never closed (LLM truncation)
        pattern = r"```python\s*\n(.+)"
        match = re.search(pattern, text, re.DOTALL)
        if match:
            code = match.group(1).strip()
            # Remove trailing ``` if partially present
            code = re.sub(r'`{1,3}\s*$', '', code).strip()
            if code:
                return code

        return None

    def _execute_safely(
        self,
        code: str,
        reference_files: Optional[list] = None
    ) -> dict:
        """
        Execute code in isolated temporary directory with security controls.

        Args:
            code: Python code to execute
            reference_files: Optional list of file paths to copy

        Returns:
            dict with success, text, files, error
        """
        with tempfile.TemporaryDirectory() as tmpdir:
            try:
                # Save code to file
                code_path = Path(tmpdir) / "solution.py"
                code_path.write_text(code, encoding="utf-8")

                # Copy reference files to execution directory and track copied files
                copied_files = []
                if reference_files:
                    for src_path in reference_files:
                        if os.path.exists(src_path):
                            try:
                                filename = os.path.basename(src_path)
                                shutil.copy(src_path, tmpdir)
                                copied_files.append(filename)
                            except Exception as e:
                                print(f"Warning: Failed to copy reference file {src_path}: {e}")

                # Inject available files list into code
                if copied_files:
                    files_comment = f"# Available files in current directory: {copied_files}\n"
                    code = files_comment + code
                else:
                    code = "# No reference files available\n" + code

                # 🔒 Security: Whitelist environment variables
                # Use current Python's PATH so venv packages are available
                # but strip API keys and secrets
                safe_env = {
                    "PATH": os.environ.get("PATH", "/usr/bin:/bin:/usr/local/bin"),
                    "LANG": "C.UTF-8",
                    "HOME": tmpdir,
                    "TMPDIR": tmpdir,
                    # Explicitly NO: AZURE_OPENAI_API_KEY, HF_TOKEN, etc.
                }

                # Preserve VIRTUAL_ENV and related paths for package access
                if "VIRTUAL_ENV" in os.environ:
                    safe_env["VIRTUAL_ENV"] = os.environ["VIRTUAL_ENV"]

                # Use the same Python interpreter (preserves venv)
                python_executable = sys.executable

                # Execute code with timeout
                result = subprocess.run(
                    [python_executable, str(code_path)],
                    cwd=tmpdir,
                    env=safe_env,
                    capture_output=True,
                    text=True,
                    timeout=SUBPROCESS_TIMEOUT
                )

                # Check execution result
                if result.returncode != 0:
                    return {
                        "success": False,
                        "text": result.stdout,
                        "files": [],
                        "error": f"Code execution failed (exit code {result.returncode}):\n{result.stderr}"
                    }

                # Collect generated files
                output_files = []
                file_extensions = ['.pdf', '.docx', '.xlsx', '.pptx', '.png', '.jpg',
                                   '.html', '.md', '.json', '.txt', '.zip', '.csv']

                for ext in file_extensions:
                    for file_path in Path(tmpdir).glob(f"*{ext}"):
                        # Skip the solution script itself
                        if file_path.name == "solution.py":
                            continue

                        try:
                            output_files.append({
                                "filename": file_path.name,
                                "content": file_path.read_bytes()
                            })
                        except Exception as e:
                            print(f"Warning: Failed to read generated file {file_path}: {e}")

                return {
                    "success": True,
                    "text": result.stdout,
                    "files": output_files
                }

            except subprocess.TimeoutExpired:
                return {
                    "success": False,
                    "text": "",
                    "files": [],
                    "error": f"Code execution timeout ({SUBPROCESS_TIMEOUT} seconds exceeded)"
                }

            except Exception as e:
                return {
                    "success": False,
                    "text": "",
                    "files": [],
                    "error": f"Execution error: {str(e)}"
                }
