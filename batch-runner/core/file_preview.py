"""
File Preview Generator — Create text previews of reference files for LLM context.

Subprocess mode cannot upload files to a sandbox like Code Interpreter.
Instead, this module reads reference files and generates concise text previews
that are injected into the LLM prompt so it can understand file contents before
generating code.

Supported formats:
- Excel (.xlsx, .xls)  — sheet names, columns, first N rows, row/column counts
- CSV (.csv)           — columns, first N rows, row count
- Word (.docx)         — full text extraction
- PDF (.pdf)           — text extraction via pdfplumber
- Text (.txt, .md, .json, .html, .xml, .yaml, .yml) — raw content
- Images (.png, .jpg, .jpeg, .webp) — dimensions and file size only
- Other formats        — file size and type info only

Usage:
    from core.file_preview import generate_all_previews

    previews = generate_all_previews(
        ["/path/to/data.xlsx", "/path/to/brief.pdf"],
        max_rows=10,
        max_chars=3000,
    )
    # Returns a formatted string block ready for prompt injection
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Tuple


# Limits
MAX_PREVIEW_ROWS = 10
MAX_PREVIEW_CHARS_PER_FILE = 3000
MAX_TOTAL_PREVIEW_CHARS = 10000

#: The longest a file name may be on the filesystems this runs on (POSIX
#: NAME_MAX). Several of the previews below put the name in a header built
#: *after* the character cap is applied — ``_preview_docx`` and its siblings
#: cut the text to ``max_chars - 200`` and then add lines around it — so the
#: name is the amount by which one preview can exceed its own cap.
MAX_FILE_NAME_CHARACTERS = 255

#: What ``generate_all_previews`` wraps around the previews it returns: the
#: opening banner, the closing line, and the blank lines between entries. These
#: sit outside the ``max_total_chars`` running total, which counts only the
#: previews themselves, so they are counted here instead of being ignored.
PREVIEW_BLOCK_WRAPPER_CHARACTERS = 300


@dataclass(frozen=True)
class ReferenceFilePromptBudget:
    """The most one reference file can add to a prompt, and what is unbounded.

    ``capped_characters`` is the part this module really does limit, worked out
    from the constants above rather than copied anywhere. ``uncapped_sections``
    names the sections that have no limit in this module at all — for those, no
    number is the truth, and saying so is more use than a figure that looks
    checked.
    """

    capped_characters: int
    uncapped_sections: Tuple[str, ...]

    @property
    def is_fully_capped(self) -> bool:
        return not self.uncapped_sections


#: The prompt sections this module supplies the text for, named as
#: core/prompt_sections.py routes them. Anything outside this set is refused
#: rather than priced at nothing.
SECTIONS_THIS_MODULE_FILLS = frozenset(
    {"file_structure", "previews", "available_files"}
)


def _characters_one_file_may_add(section: str) -> Optional[int]:
    """How much one reference file may add through ``section``, or ``None``.

    ``None`` means this module sets no limit on that section. It is a
    different answer from zero and must not be added up as though it were.

    Worked out from the constants above on every call rather than once at
    import, so raising a cap moves this — and moves everything that holds a
    figure against it — instead of leaving a stale number behind.

    Raises:
        KeyError: for a section this module does not fill. Guessing zero for
                  one would quietly lower a cost ceiling.
    """
    if section not in SECTIONS_THIS_MODULE_FILLS:
        raise KeyError(section)
    if section == "file_structure":
        # build_file_structure_info writes one line per sheet listing every
        # column header it finds. Nothing here cuts that off, and a workbook
        # may carry any number of columns, so there is no ceiling to read.
        return None
    if section == "previews":
        # One preview, cut at MAX_PREVIEW_CHARS_PER_FILE, plus the file name
        # in the header that is written after the cut, plus this file's share
        # of the block wrapper.
        return (
            MAX_PREVIEW_CHARS_PER_FILE
            + MAX_FILE_NAME_CHARACTERS
            + PREVIEW_BLOCK_WRAPPER_CHARACTERS
        )
    # available_files: this file's name in the list, with the quotes around it
    # and the separator that follows.
    return MAX_FILE_NAME_CHARACTERS + len("', '")


def reference_file_prompt_budget(
    sections: Iterable[str],
) -> ReferenceFilePromptBudget:
    """Work out what one reference file can add through the sections named.

    Args:
        sections: prompt section ids, as core/prompt_sections.py knows them.

    Raises:
        KeyError: if a section is named that this module does not fill.
    """
    capped = 0
    uncapped: List[str] = []
    for section in sections:
        per_file = _characters_one_file_may_add(section)
        if per_file is None:
            uncapped.append(section)
        else:
            capped += per_file
    return ReferenceFilePromptBudget(
        capped_characters=capped,
        uncapped_sections=tuple(sorted(uncapped)),
    )


def generate_file_preview(
    file_path: str,
    max_rows: int = MAX_PREVIEW_ROWS,
    max_chars: int = MAX_PREVIEW_CHARS_PER_FILE,
) -> str:
    """
    Generate a text preview of a single reference file.

    Args:
        file_path: Absolute path to the file
        max_rows: Max rows to show for tabular data
        max_chars: Max characters for text content

    Returns:
        Formatted string preview of the file
    """
    path = Path(file_path)
    if not path.exists():
        return f"[File not found: {path.name}]"

    ext = path.suffix.lower()
    filename = path.name
    file_size = path.stat().st_size

    try:
        if ext in (".xlsx", ".xls"):
            return _preview_excel(path, filename, file_size, max_rows, max_chars)
        elif ext == ".csv":
            return _preview_csv(path, filename, file_size, max_rows, max_chars)
        elif ext == ".docx":
            return _preview_docx(path, filename, file_size, max_chars)
        elif ext == ".pdf":
            return _preview_pdf(path, filename, file_size, max_chars)
        elif ext in (".txt", ".md", ".json", ".html", ".xml", ".yaml", ".yml"):
            return _preview_text(path, filename, file_size, max_chars)
        elif ext in (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"):
            return _preview_image(path, filename, file_size)
        else:
            return (
                f"📎 {filename} ({_fmt_size(file_size)})\n"
                f"  Type: {ext} — preview not supported, file is copied to working directory"
            )
    except Exception as e:
        return f"📎 {filename} ({_fmt_size(file_size)})\n  ⚠️ Preview error: {str(e)}"


def generate_all_previews(
    file_paths: List[str],
    max_rows: int = MAX_PREVIEW_ROWS,
    max_chars: int = MAX_PREVIEW_CHARS_PER_FILE,
    max_total_chars: int = MAX_TOTAL_PREVIEW_CHARS,
) -> Optional[str]:
    """
    Generate combined previews for all reference files.

    Args:
        file_paths: List of absolute file paths
        max_rows: Max rows per tabular file
        max_chars: Max characters per file preview
        max_total_chars: Max total characters for all previews combined

    Returns:
        Combined preview string, or None if no files
    """
    if not file_paths:
        return None

    previews = []
    total_chars = 0

    for fp in file_paths:
        preview = generate_file_preview(fp, max_rows, max_chars)
        if total_chars + len(preview) > max_total_chars:
            remaining = max_total_chars - total_chars
            if remaining > 100:
                preview = preview[:remaining] + "\n  ... (truncated)"
            else:
                previews.append(f"  ... ({len(file_paths) - len(previews)} more files, truncated)")
                break
        previews.append(preview)
        total_chars += len(preview)

    header = (
        "═══ REFERENCE FILES PREVIEW ═══\n"
        "The following reference files are available in the current directory.\n"
        "Use EXACT column names, sheet names, and data values shown below.\n"
    )
    return header + "\n\n" + "\n\n".join(previews) + "\n═══ END REFERENCE FILES ═══"


# ─── Format helpers ───


def _fmt_size(size_bytes: int) -> str:
    """Format file size as human-readable string."""
    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"


# ─── Format-specific preview generators ───


def _preview_excel(
    path: Path, filename: str, file_size: int, max_rows: int, max_chars: int
) -> str:
    """Preview Excel file: sheet names, columns, first N rows."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lines = [f"📊 {filename} ({_fmt_size(file_size)})"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            lines.append(f"  Sheet: {sheet_name} — (empty)")
            continue

        # First row as headers
        headers = [str(h) if h is not None else "" for h in rows[0]]
        data_rows = rows[1:]
        total_rows = len(data_rows)

        lines.append(f"  Sheet: {sheet_name} ({total_rows} rows × {len(headers)} cols)")
        lines.append(f"  Columns: {', '.join(headers)}")

        # Show first N data rows
        show_rows = data_rows[:max_rows]
        if show_rows:
            lines.append(f"  First {len(show_rows)} rows:")
            for i, row in enumerate(show_rows):
                row_str = " | ".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                lines.append(f"    [{i+1}] {row_str}")

        if total_rows > max_rows:
            lines.append(f"    ... ({total_rows - max_rows} more rows)")

    wb.close()
    result = "\n".join(lines)
    return result[:max_chars] if len(result) > max_chars else result


def _preview_csv(
    path: Path, filename: str, file_size: int, max_rows: int, max_chars: int
) -> str:
    """Preview CSV file: columns, first N rows."""
    import csv

    lines = [f"📊 {filename} ({_fmt_size(file_size)})"]

    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        rows = []
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= max_rows + 1:  # +1 for header
                break

    if not rows:
        lines.append("  (empty file)")
        return "\n".join(lines)

    headers = rows[0]
    data_rows = rows[1:]

    lines.append(f"  Columns: {', '.join(headers)}")
    lines.append(f"  First {len(data_rows)} rows:")
    for i, row in enumerate(data_rows):
        row_str = " | ".join(row)
        lines.append(f"    [{i+1}] {row_str}")

    result = "\n".join(lines)
    return result[:max_chars] if len(result) > max_chars else result


def _preview_docx(path: Path, filename: str, file_size: int, max_chars: int) -> str:
    """Preview Word document: extract text content."""
    from docx import Document

    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

    lines = [f"📝 {filename} ({_fmt_size(file_size)}, {len(paragraphs)} paragraphs)"]

    text = "\n".join(paragraphs)
    if len(text) > max_chars - 200:
        text = text[: max_chars - 200] + "\n  ... (truncated)"

    lines.append(text)

    return "\n".join(lines)


def _preview_pdf(path: Path, filename: str, file_size: int, max_chars: int) -> str:
    """Preview PDF: extract text via pdfplumber."""
    import pdfplumber

    lines = [f"📄 {filename} ({_fmt_size(file_size)})"]

    with pdfplumber.open(path) as pdf:
        num_pages = len(pdf.pages)
        lines.append(f"  Pages: {num_pages}")

        text_parts = []
        for page in pdf.pages[:5]:  # First 5 pages max
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)

        text = "\n".join(text_parts)
        if len(text) > max_chars - 200:
            text = text[: max_chars - 200] + "\n  ... (truncated)"

        if text.strip():
            lines.append(text)
        else:
            lines.append("  (no extractable text — may be image-based PDF)")

    return "\n".join(lines)


def _preview_text(path: Path, filename: str, file_size: int, max_chars: int) -> str:
    """Preview text-based file."""
    lines = [f"📝 {filename} ({_fmt_size(file_size)})"]

    try:
        content = path.read_text(encoding="utf-8", errors="replace")
        if len(content) > max_chars - 100:
            content = content[: max_chars - 100] + "\n  ... (truncated)"
        lines.append(content)
    except Exception as e:
        lines.append(f"  ⚠️ Read error: {e}")

    return "\n".join(lines)


def _preview_image(path: Path, filename: str, file_size: int) -> str:
    """Preview image: dimensions and file size only."""
    info = f"🖼️ {filename} ({_fmt_size(file_size)})"

    try:
        from PIL import Image

        with Image.open(path) as img:
            w, h = img.size
            info += f"\n  Dimensions: {w}×{h}px, Mode: {img.mode}"
    except Exception:
        info += "\n  (could not read image dimensions)"

    return info


# ─── Structure info (token-efficient metadata only) ───


def build_file_structure_info(reference_files: list) -> str:
    """Read reference file structure and return a prompt-ready summary.

    Reads column names and shape from each reference file.
    Supports: .xlsx, .xls, .csv, .tsv, .parquet, .json

    Token-efficient: metadata only, no row content.
    Returns empty string if no files provided or all fail to read.

    Used by both subprocess_runner.py and code_interpreter.py to prevent
    LLM from hardcoding column names without inspecting the actual file.
    """
    if not reference_files:
        return ""

    lines = ["## Reference File Structure (auto-detected)"]
    any_success = False

    for fpath in reference_files:
        path = Path(fpath)
        if not path.exists():
            continue
        suffix = path.suffix.lower()

        try:
            if suffix in (".xlsx", ".xls"):
                import openpyxl
                wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                    headers = [str(h) for h in headers if h is not None]
                    nrows = (ws.max_row or 1) - 1
                    label = (
                        f"{path.name} (sheet: {sheet_name})"
                        if len(wb.sheetnames) > 1
                        else path.name
                    )
                    lines.append(f"\n{label}: ~{nrows} rows × {len(headers)} cols")
                    lines.append(f"  columns: {headers}")
                wb.close()
                any_success = True

            elif suffix in (".csv", ".tsv"):
                import csv
                sep = "\t" if suffix == ".tsv" else ","
                with open(fpath, newline="", encoding="utf-8-sig") as f:
                    reader = csv.reader(f, delimiter=sep)
                    headers = next(reader)
                    nrows = sum(1 for _ in reader)
                lines.append(f"\n{path.name}: ~{nrows} rows × {len(headers)} cols")
                lines.append(f"  columns: {headers}")
                any_success = True

            elif suffix == ".parquet":
                import pyarrow.parquet as pq
                schema = pq.read_schema(fpath)
                meta = pq.read_metadata(fpath)
                lines.append(f"\n{path.name}: {meta.num_rows} rows × {len(schema.names)} cols")
                lines.append(f"  columns: {schema.names}")
                any_success = True

            elif suffix == ".json":
                import json as _json
                with open(fpath, encoding="utf-8") as f:
                    data = _json.load(f)
                if isinstance(data, list) and data:
                    keys = list(data[0].keys()) if isinstance(data[0], dict) else []
                    lines.append(f"\n{path.name}: list of {len(data)} records")
                    if keys:
                        lines.append(f"  keys: {keys}")
                elif isinstance(data, dict):
                    lines.append(f"\n{path.name}: dict with keys: {list(data.keys())[:20]}")
                any_success = True

        except Exception:
            # 읽기 실패 시 조용히 스킵 — 추론 크래시 방지
            continue

    return "\n".join(lines) if any_success else ""
